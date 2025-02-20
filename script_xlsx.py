#взаимодействие с xlsx. Установить библиотеку: pip install openpyxl
from openpyxl import load_workbook

workbook = load_workbook("tmp/СВОЙ ФАЙЛ.xlsx")
sheet = workbook.active
print(sheet.cell(row=1, column=2).value)